"""
build_sla_dashboard.py
------------------------
Builds a formula-driven Excel SLA/operational-analytics dashboard
(dashboard.xlsx) from tickets.csv. All KPIs are computed via native Excel
formulas (COUNTIF/COUNTIFS/AVERAGEIF/AVERAGE) referencing the raw data
sheet — not pre-computed static values — so the workbook recalculates
correctly if the underlying dataset changes.

Output: ../dashboard.xlsx
    Sheet "Raw Data"  — full ticket dataset
    Sheet "Dashboard" — KPI summary, category/priority/agent breakdowns,
                        native BarChart/PieChart objects, conditional
                        formatting on the SLA Compliance KPI (<85% -> red)

Usage:
    python3 build_sla_dashboard.py
"""
import csv
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

DATA_PATH = "../data/tickets.csv"
OUTPUT_PATH = "../dashboard.xlsx"
SLA_TARGET_THRESHOLD = 0.85  # conditional-formatting breakpoint

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True, color="1F4E79")
KPI_FONT = Font(size=22, bold=True, color="1F4E79")
KPI_LABEL_FONT = Font(size=10, bold=True, color="595959")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUMERIC_FIELDS = {"resolution_time_hrs", "sla_target_hrs", "csat_score"}


def load_rows():
    with open(DATA_PATH) as f:
        return list(csv.DictReader(f))


def style_header(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER


def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def write_raw_data_sheet(wb, rows):
    """Writes the source dataset, casting numeric fields to actual numeric
    types (not strings) so downstream Excel formulas (AVERAGE, AVERAGEIF)
    evaluate correctly rather than returning #DIV/0!."""
    ws = wb.active
    ws.title = "Raw Data"
    headers = list(rows[0].keys())
    ws.append(headers)
    style_header(ws, 1, 1, len(headers))
    for r in rows:
        ws.append([float(r[h]) if h in NUMERIC_FIELDS else r[h] for h in headers])
    autosize(ws, {get_column_letter(i + 1): 16 for i in range(len(headers))})
    return ws, len(rows) + 1  # last_row (1-indexed, includes header)


def write_kpi_summary(ws, last_row):
    ws["B2"] = "NorthBridge Retail Ltd — IT Service Desk Performance Dashboard"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = "Data window: Jun 1 – Jul 10, 2026  |  Source: Raw Data sheet (live formulas)"
    ws["B3"].font = Font(italic=True, color="808080", size=10)

    kpis = [
        ("Total Tickets", f"=COUNTA('Raw Data'!A2:A{last_row})", "B5"),
        ("SLA Compliance", f"=COUNTIF('Raw Data'!K2:K{last_row},\"Yes\")/COUNTA('Raw Data'!K2:K{last_row})", "D5"),
        ("First Contact Resolution", f"=COUNTIF('Raw Data'!L2:L{last_row},\"Yes\")/COUNTA('Raw Data'!L2:L{last_row})", "F5"),
        ("Avg Resolution Time (hrs)", f"=AVERAGE('Raw Data'!I2:I{last_row})", "H5"),
        ("Reopen Rate", f"=COUNTIF('Raw Data'!M2:M{last_row},\"Yes\")/COUNTA('Raw Data'!M2:M{last_row})", "J5"),
        ("Avg CSAT (/5)", f"=AVERAGE('Raw Data'!N2:N{last_row})", "L5"),
    ]
    cols = ["B", "D", "F", "H", "J", "L"]
    for (label, _, cell), col in zip(kpis, cols):
        ws[f"{col}4"] = label
        ws[f"{col}4"].font = KPI_LABEL_FONT
    for label, formula, cell in kpis:
        ws[cell] = formula
        ws[cell].font = KPI_FONT
    ws["D5"].number_format = "0.0%"
    ws["F5"].number_format = "0.0%"
    ws["H5"].number_format = "0.00"
    ws["J5"].number_format = "0.0%"
    ws["L5"].number_format = "0.00"

    ws.conditional_formatting.add(
        "D5",
        CellIsRule(operator="lessThan", formula=[str(SLA_TARGET_THRESHOLD)],
                   fill=PatternFill("solid", fgColor="FFC7CE")),
    )


def write_category_breakdown(ws, rows, last_row):
    ws["B8"] = "Tickets & SLA Compliance by Category"
    ws["B8"].font = Font(bold=True, size=12, color="1F4E79")
    header_row = 9
    for col, label in zip((2, 3, 4, 5), ("Category", "Ticket Count", "SLA Compliance %", "Avg Resolution (hrs)")):
        ws.cell(row=header_row, column=col, value=label)
    style_header(ws, header_row, 2, 5)

    categories = sorted(set(r["category"] for r in rows))
    for i, cat in enumerate(categories):
        r = header_row + 1 + i
        ws.cell(row=r, column=2, value=cat)
        ws.cell(row=r, column=3, value=f"=COUNTIF('Raw Data'!B2:B{last_row},B{r})")
        ws.cell(row=r, column=4,
                value=f"=COUNTIFS('Raw Data'!B2:B{last_row},B{r},'Raw Data'!K2:K{last_row},\"Yes\")/C{r}")
        ws.cell(row=r, column=4).number_format = "0.0%"
        ws.cell(row=r, column=5,
                value=f"=AVERAGEIF('Raw Data'!B2:B{last_row},B{r},'Raw Data'!I2:I{last_row})")
        ws.cell(row=r, column=5).number_format = "0.00"
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = BORDER
    return header_row, header_row + len(categories)


def write_priority_breakdown(ws, last_row):
    ws["G8"] = "Tickets by Priority"
    ws["G8"].font = Font(bold=True, size=12, color="1F4E79")
    header_row = 9
    ws.cell(row=header_row, column=7, value="Priority")
    ws.cell(row=header_row, column=8, value="Ticket Count")
    style_header(ws, header_row, 7, 8)
    priorities = ["P1 - Critical", "P2 - High", "P3 - Moderate", "P4 - Low"]
    for i, pri in enumerate(priorities):
        r = header_row + 1 + i
        ws.cell(row=r, column=7, value=pri)
        ws.cell(row=r, column=8, value=f"=COUNTIF('Raw Data'!D2:D{last_row},G{r})")
        for c in range(7, 9):
            ws.cell(row=r, column=c).border = BORDER
    return header_row, header_row + len(priorities)


def write_agent_breakdown(ws, rows, last_row):
    ws["B16"] = "Agent Performance"
    ws["B16"].font = Font(bold=True, size=12, color="1F4E79")
    header_row = 17
    for col, label in zip((2, 3, 4, 5), ("Agent", "Tickets Handled", "SLA Compliance %", "Avg CSAT")):
        ws.cell(row=header_row, column=col, value=label)
    style_header(ws, header_row, 2, 5)
    agents = sorted(set(r["agent"] for r in rows))
    for i, ag in enumerate(agents):
        r = header_row + 1 + i
        ws.cell(row=r, column=2, value=ag)
        ws.cell(row=r, column=3, value=f"=COUNTIF('Raw Data'!F2:F{last_row},B{r})")
        ws.cell(row=r, column=4,
                value=f"=COUNTIFS('Raw Data'!F2:F{last_row},B{r},'Raw Data'!K2:K{last_row},\"Yes\")/C{r}")
        ws.cell(row=r, column=4).number_format = "0.0%"
        ws.cell(row=r, column=5, value=f"=AVERAGEIF('Raw Data'!F2:F{last_row},B{r},'Raw Data'!N2:N{last_row})")
        ws.cell(row=r, column=5).number_format = "0.00"
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = BORDER
    return header_row, header_row + len(agents)


def add_charts(ws, cat_range, pri_range, agent_range):
    cat_header_row, cat_last_row = cat_range
    pri_header_row, pri_last_row = pri_range
    ag_header_row, ag_last_row = agent_range

    bar1 = BarChart()
    bar1.title = "Ticket Volume by Category"
    bar1.y_axis.title = "Tickets"
    bar1.add_data(Reference(ws, min_col=3, min_row=cat_header_row, max_row=cat_last_row), titles_from_data=True)
    bar1.set_categories(Reference(ws, min_col=2, min_row=cat_header_row + 1, max_row=cat_last_row))
    bar1.width, bar1.height = 16, 9
    ws.add_chart(bar1, "B25")

    pie1 = PieChart()
    pie1.title = "Tickets by Priority"
    pie1.add_data(Reference(ws, min_col=8, min_row=pri_header_row, max_row=pri_last_row), titles_from_data=True)
    pie1.set_categories(Reference(ws, min_col=7, min_row=pri_header_row + 1, max_row=pri_last_row))
    pie1.width, pie1.height = 12, 9
    ws.add_chart(pie1, "J25")

    bar2 = BarChart()
    bar2.type = "col"
    bar2.title = "SLA Compliance % by Agent"
    bar2.add_data(Reference(ws, min_col=4, min_row=ag_header_row, max_row=ag_last_row), titles_from_data=True)
    bar2.set_categories(Reference(ws, min_col=2, min_row=ag_header_row + 1, max_row=ag_last_row))
    bar2.width, bar2.height = 16, 9
    ws.add_chart(bar2, "B43")


def main():
    rows = load_rows()
    wb = Workbook()

    ws_data, last_row = write_raw_data_sheet(wb, rows)

    ws = wb.create_sheet("Dashboard")
    write_kpi_summary(ws, last_row)
    cat_range = write_category_breakdown(ws, rows, last_row)
    pri_range = write_priority_breakdown(ws, last_row)
    agent_range = write_agent_breakdown(ws, rows, last_row)
    add_charts(ws, cat_range, pri_range, agent_range)

    autosize(ws, {"B": 26, "C": 16, "D": 18, "E": 20, "F": 14, "G": 16, "H": 14, "J": 14, "L": 12})

    wb.save(OUTPUT_PATH)
    print(f"Dashboard built -> {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
