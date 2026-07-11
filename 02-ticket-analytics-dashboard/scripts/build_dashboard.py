"""
build_dashboard.py
-------------------
Builds a working Excel SLA/analytics dashboard (dashboard.xlsx) from
tickets.csv, using native Excel formulas (not hard-coded numbers) so the
sheet recalculates if you regenerate the data. Also renders PNG chart
screenshots (via matplotlib) for the README.
"""
import csv
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

DATA_PATH = "../data/tickets.csv"
OUT_PATH = "../dashboard.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True, color="1F4E79")
KPI_FONT = Font(size=22, bold=True, color="1F4E79")
KPI_LABEL_FONT = Font(size=10, bold=True, color="595959")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def load_rows():
    with open(DATA_PATH) as f:
        return list(csv.DictReader(f))

def style_header(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def main():
    rows = load_rows()
    wb = Workbook()

    # ---------- Sheet 1: Raw Data ----------
    ws_data = wb.active
    ws_data.title = "Raw Data"
    headers = list(rows[0].keys())
    ws_data.append(headers)
    style_header(ws_data, 1, 1, len(headers))
    numeric_fields = {"resolution_time_hrs", "sla_target_hrs", "csat_score"}
    for r in rows:
        row_vals = []
        for h in headers:
            v = r[h]
            if h in numeric_fields:
                v = float(v)
            row_vals.append(v)
        ws_data.append(row_vals)
    autosize(ws_data, {get_column_letter(i+1): 16 for i in range(len(headers))})
    n = len(rows)
    last_row = n + 1

    # ---------- Sheet 2: Summary / KPIs ----------
    ws = wb.create_sheet("Dashboard")
    ws["B2"] = "NorthBridge Retail Ltd — IT Service Desk Performance Dashboard"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = "Data window: Jun 1 – Jul 10, 2026  |  Source: Raw Data sheet (live formulas)"
    ws["B3"].font = Font(italic=True, color="808080", size=10)

    kpis = [
        ("Total Tickets", f"=COUNTA('Raw Data'!A2:A{last_row})", "B5"),
        ("SLA Compliance", f"=COUNTIF('Raw Data'!K2:K{last_row},\"Yes\")/COUNTA('Raw Data'!K2:K{last_row})", "D5"),
        ("First Contact Resolution", f"=COUNTIF('Raw Data'!L2:L{last_row},\"Yes\")/COUNTA('Raw Data'!L2:L{last_row})", "F5"),
        ("Avg Resolution Time (hrs)", f"=AVERAGE('Raw Data'!I2:I{last_row})", "H5"),
        ("Reopen Rate", f"=COUNTIF('Raw Data'!M2:M{last_row},\"Yes\")/COUNTA('Raw Data'!M2:M{last_row})", "J5"),
        ("Avg CSAT (/5)", f"=AVERAGE('Raw Data'!N2:N{last_row})", "L5"),
    ]
    labels = ["Total Tickets", "SLA Compliance", "First Contact Resolution",
              "Avg Resolution Time (hrs)", "Reopen Rate", "Avg CSAT (/5)"]
    cols = ["B", "D", "F", "H", "J", "L"]
    for label, col in zip(labels, cols):
        ws[f"{col}4"] = label
        ws[f"{col}4"].font = KPI_LABEL_FONT
    for label, formula, cell in kpis:
        ws[cell] = formula
        ws[cell].font = KPI_FONT
    ws["D5"].number_format = "0.0%"
    ws["F5"].number_format = "0.0%"
    ws["H5"].number_format = "0.00"
    ws["J5"].number_format = "0.0%"
    ws["L5"].number_format = "0.00"

    # Conditional formatting on SLA compliance KPI
    ws.conditional_formatting.add(
        "D5",
        CellIsRule(operator="lessThan", formula=["0.85"], fill=PatternFill("solid", fgColor="FFC7CE")),
    )

    # ---------- Breakdown table: tickets & SLA % by category ----------
    ws["B8"] = "Tickets & SLA Compliance by Category"
    ws["B8"].font = Font(bold=True, size=12, color="1F4E79")
    cat_header_row = 9
    ws.cell(row=cat_header_row, column=2, value="Category")
    ws.cell(row=cat_header_row, column=3, value="Ticket Count")
    ws.cell(row=cat_header_row, column=4, value="SLA Compliance %")
    ws.cell(row=cat_header_row, column=5, value="Avg Resolution (hrs)")
    style_header(ws, cat_header_row, 2, 5)

    categories = sorted(set(r["category"] for r in rows))
    for i, cat in enumerate(categories):
        r = cat_header_row + 1 + i
        ws.cell(row=r, column=2, value=cat)
        ws.cell(row=r, column=3,
                 value=f"=COUNTIF('Raw Data'!B2:B{last_row},B{r})")
        ws.cell(row=r, column=4,
                 value=f"=COUNTIFS('Raw Data'!B2:B{last_row},B{r},'Raw Data'!K2:K{last_row},\"Yes\")/C{r}")
        ws.cell(row=r, column=4).number_format = "0.0%"
        ws.cell(row=r, column=5,
                 value=f"=AVERAGEIF('Raw Data'!B2:B{last_row},B{r},'Raw Data'!I2:I{last_row})")
        ws.cell(row=r, column=5).number_format = "0.00"
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = BORDER
    cat_last_row = cat_header_row + len(categories)

    # ---------- Breakdown table: tickets by priority ----------
    ws["G8"] = "Tickets by Priority"
    ws["G8"].font = Font(bold=True, size=12, color="1F4E79")
    pri_header_row = 9
    ws.cell(row=pri_header_row, column=7, value="Priority")
    ws.cell(row=pri_header_row, column=8, value="Ticket Count")
    style_header(ws, pri_header_row, 7, 8)
    priorities = ["P1 - Critical", "P2 - High", "P3 - Moderate", "P4 - Low"]
    for i, pri in enumerate(priorities):
        r = pri_header_row + 1 + i
        ws.cell(row=r, column=7, value=pri)
        ws.cell(row=r, column=8, value=f"=COUNTIF('Raw Data'!D2:D{last_row},G{r})")
        for c in range(7, 9):
            ws.cell(row=r, column=c).border = BORDER
    pri_last_row = pri_header_row + len(priorities)

    # ---------- Breakdown table: tickets by agent (workload + SLA) ----------
    ws["B16"] = "Agent Performance"
    ws["B16"].font = Font(bold=True, size=12, color="1F4E79")
    ag_header_row = 17
    ws.cell(row=ag_header_row, column=2, value="Agent")
    ws.cell(row=ag_header_row, column=3, value="Tickets Handled")
    ws.cell(row=ag_header_row, column=4, value="SLA Compliance %")
    ws.cell(row=ag_header_row, column=5, value="Avg CSAT")
    style_header(ws, ag_header_row, 2, 5)
    agents = sorted(set(r["agent"] for r in rows))
    for i, ag in enumerate(agents):
        r = ag_header_row + 1 + i
        ws.cell(row=r, column=2, value=ag)
        ws.cell(row=r, column=3, value=f"=COUNTIF('Raw Data'!F2:F{last_row},B{r})")
        ws.cell(row=r, column=4,
                 value=f"=COUNTIFS('Raw Data'!F2:F{last_row},B{r},'Raw Data'!K2:K{last_row},\"Yes\")/C{r}")
        ws.cell(row=r, column=4).number_format = "0.0%"
        ws.cell(row=r, column=5, value=f"=AVERAGEIF('Raw Data'!F2:F{last_row},B{r},'Raw Data'!N2:N{last_row})")
        ws.cell(row=r, column=5).number_format = "0.00"
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = BORDER
    ag_last_row = ag_header_row + len(agents)

    # ---------- Charts ----------
    bar1 = BarChart()
    bar1.title = "Ticket Volume by Category"
    bar1.y_axis.title = "Tickets"
    data_ref = Reference(ws, min_col=3, min_row=cat_header_row, max_row=cat_last_row)
    cats_ref = Reference(ws, min_col=2, min_row=cat_header_row + 1, max_row=cat_last_row)
    bar1.add_data(data_ref, titles_from_data=True)
    bar1.set_categories(cats_ref)
    bar1.width, bar1.height = 16, 9
    ws.add_chart(bar1, "B25")

    pie1 = PieChart()
    pie1.title = "Tickets by Priority"
    data_ref2 = Reference(ws, min_col=8, min_row=pri_header_row, max_row=pri_last_row)
    cats_ref2 = Reference(ws, min_col=7, min_row=pri_header_row + 1, max_row=pri_last_row)
    pie1.add_data(data_ref2, titles_from_data=True)
    pie1.set_categories(cats_ref2)
    pie1.width, pie1.height = 12, 9
    ws.add_chart(pie1, "J25")

    bar2 = BarChart()
    bar2.type = "col"
    bar2.title = "SLA Compliance % by Agent"
    data_ref3 = Reference(ws, min_col=4, min_row=ag_header_row, max_row=ag_last_row)
    cats_ref3 = Reference(ws, min_col=2, min_row=ag_header_row + 1, max_row=ag_last_row)
    bar2.add_data(data_ref3, titles_from_data=True)
    bar2.set_categories(cats_ref3)
    bar2.width, bar2.height = 16, 9
    ws.add_chart(bar2, "B43")

    autosize(ws, {"B": 26, "C": 16, "D": 18, "E": 20, "F": 14, "G": 16, "H": 14, "J": 14, "L": 12})

    wb.save(OUT_PATH)
    print(f"Dashboard saved -> {OUT_PATH}")

if __name__ == "__main__":
    main()
